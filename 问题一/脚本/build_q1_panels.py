from __future__ import annotations

import csv
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


PROBLEM_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = PROBLEM_DIR.parent / "2023年C题"
OUT_DIR = PROBLEM_DIR / "结果"
CATALOG_PATH = SOURCE_DIR / "附件1.xlsx"
FLOW_PATH = SOURCE_DIR / "附件2.xlsx"


def as_text(value) -> str:
    return "" if value is None else str(value).strip()


def day_from_value(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise TypeError(f"invalid date value: {value!r}")


def add_row(bucket: dict, quantity: float, sale_type: str) -> None:
    bucket["transaction_rows"] += 1
    bucket["net_sales_kg"] += quantity
    if sale_type == "销售":
        bucket["sales_rows"] += 1
        bucket["sales_qty_kg"] += quantity
    elif sale_type == "退货":
        bucket["return_rows"] += 1
        bucket["return_qty_kg"] += quantity
    else:
        raise ValueError(f"unexpected sales type: {sale_type!r}")


def write_csv(path: Path, fieldnames: list[str], rows) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def rounded(value: float) -> float:
    return round(float(value), 6)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    catalog: list[dict[str, str]] = []
    catalog_by_code: dict[str, dict[str, str]] = {}
    catalog_workbook = load_workbook(CATALOG_PATH, read_only=True, data_only=True)
    worksheet = catalog_workbook.active
    headers = [as_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    expected = ["单品编码", "单品名称", "分类编码", "分类名称"]
    if headers != expected:
        raise ValueError(f"unexpected attachment 1 headers: {headers!r}")
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        product = {
            "product_code": as_text(row[0]),
            "product_name": as_text(row[1]),
            "category_code": as_text(row[2]),
            "category_name": as_text(row[3]),
        }
        if not product["product_code"] or product["product_code"] in catalog_by_code:
            raise ValueError(f"invalid or duplicate catalog code: {product!r}")
        catalog.append(product)
        catalog_by_code[product["product_code"]] = product
    catalog_workbook.close()

    flow_agg: dict[tuple[date, str], dict[str, float | int]] = defaultdict(
        lambda: {
            "sales_qty_kg": 0.0,
            "return_qty_kg": 0.0,
            "net_sales_kg": 0.0,
            "transaction_rows": 0,
            "sales_rows": 0,
            "return_rows": 0,
        }
    )
    category_agg: dict[tuple[date, str], dict[str, float | int]] = defaultdict(
        lambda: {
            "sales_qty_kg": 0.0,
            "return_qty_kg": 0.0,
            "net_sales_kg": 0.0,
            "transaction_rows": 0,
            "sales_rows": 0,
            "return_rows": 0,
        }
    )
    observed_dates: set[date] = set()
    observed_codes: set[str] = set()
    raw_rows = 0
    raw_sales_rows = 0
    raw_return_rows = 0
    raw_sales_qty = 0.0
    raw_return_qty = 0.0
    min_date: date | None = None
    max_date: date | None = None

    flow_workbook = load_workbook(FLOW_PATH, read_only=True, data_only=True)
    worksheet = flow_workbook.active
    headers = [as_text(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    expected = [
        "销售日期",
        "扫码销售时间",
        "单品编码",
        "销量(千克)",
        "销售单价(元/千克)",
        "销售类型",
        "是否打折销售",
    ]
    if headers != expected:
        raise ValueError(f"unexpected attachment 2 headers: {headers!r}")

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        raw_rows += 1
        sales_date = day_from_value(row[0])
        product_code = as_text(row[2])
        quantity = float(row[3])
        sale_type = as_text(row[5])
        if product_code not in catalog_by_code:
            raise ValueError(f"unmapped product code: {product_code!r}")
        if sale_type not in {"销售", "退货"}:
            raise ValueError(f"unexpected sales type: {sale_type!r}")
        if sale_type == "销售" and quantity <= 0:
            raise ValueError(f"non-positive sales quantity: {row!r}")
        if sale_type == "退货" and quantity >= 0:
            raise ValueError(f"non-negative return quantity: {row!r}")

        observed_dates.add(sales_date)
        observed_codes.add(product_code)
        min_date = sales_date if min_date is None else min(min_date, sales_date)
        max_date = sales_date if max_date is None else max(max_date, sales_date)

        product_key = (sales_date, product_code)
        add_row(flow_agg[product_key], quantity, sale_type)

        category_name = catalog_by_code[product_code]["category_name"]
        category_key = (sales_date, category_name)
        add_row(category_agg[category_key], quantity, sale_type)

        if sale_type == "销售":
            raw_sales_rows += 1
            raw_sales_qty += quantity
        else:
            raw_return_rows += 1
            raw_return_qty += quantity
    flow_workbook.close()

    if min_date is None or max_date is None:
        raise ValueError("no flow rows found")

    category_order: list[dict[str, str]] = []
    seen_categories: set[str] = set()
    for product in catalog:
        category_name = product["category_name"]
        if category_name not in seen_categories:
            category_order.append(
                {"category_code": product["category_code"], "category_name": category_name}
            )
            seen_categories.add(category_name)

    all_dates: list[date] = []
    cursor = min_date
    while cursor <= max_date:
        all_dates.append(cursor)
        cursor += timedelta(days=1)

    item_fields = [
        "sales_date",
        "date_has_records",
        "product_code",
        "product_name",
        "category_code",
        "category_name",
        "product_has_records",
        "sales_qty_kg",
        "return_qty_kg",
        "net_sales_kg",
        "transaction_rows",
        "sales_rows",
        "return_rows",
    ]

    def item_rows():
        for sales_date in all_dates:
            date_has_records = int(sales_date in observed_dates)
            for product in catalog:
                bucket = flow_agg.get((sales_date, product["product_code"]))
                product_has_records = int(bucket is not None)
                if not date_has_records:
                    values = {field: "" for field in item_fields}
                    values.update(
                        {
                            "sales_date": sales_date.isoformat(),
                            "date_has_records": 0,
                            "product_code": product["product_code"],
                            "product_name": product["product_name"],
                            "category_code": product["category_code"],
                            "category_name": product["category_name"],
                            "product_has_records": 0,
                        }
                    )
                    yield values
                    continue
                bucket = bucket or {
                    "sales_qty_kg": 0.0,
                    "return_qty_kg": 0.0,
                    "net_sales_kg": 0.0,
                    "transaction_rows": 0,
                    "sales_rows": 0,
                    "return_rows": 0,
                }
                yield {
                    "sales_date": sales_date.isoformat(),
                    "date_has_records": date_has_records,
                    "product_code": product["product_code"],
                    "product_name": product["product_name"],
                    "category_code": product["category_code"],
                    "category_name": product["category_name"],
                    "product_has_records": product_has_records,
                    "sales_qty_kg": rounded(bucket["sales_qty_kg"]),
                    "return_qty_kg": rounded(bucket["return_qty_kg"]),
                    "net_sales_kg": rounded(bucket["net_sales_kg"]),
                    "transaction_rows": bucket["transaction_rows"],
                    "sales_rows": bucket["sales_rows"],
                    "return_rows": bucket["return_rows"],
                }

    write_csv(OUT_DIR / "daily_item_panel.csv", item_fields, item_rows())

    category_fields = [
        "sales_date",
        "date_has_records",
        "category_code",
        "category_name",
        "category_has_records",
        "sales_qty_kg",
        "return_qty_kg",
        "net_sales_kg",
        "transaction_rows",
        "sales_rows",
        "return_rows",
        "active_product_count",
    ]

    def category_rows():
        for sales_date in all_dates:
            date_has_records = int(sales_date in observed_dates)
            for category in category_order:
                bucket = category_agg.get((sales_date, category["category_name"]))
                category_has_records = int(bucket is not None)
                if not date_has_records:
                    values = {field: "" for field in category_fields}
                    values.update(
                        {
                            "sales_date": sales_date.isoformat(),
                            "date_has_records": 0,
                            "category_code": category["category_code"],
                            "category_name": category["category_name"],
                            "category_has_records": 0,
                        }
                    )
                    yield values
                    continue
                bucket = bucket or {
                    "sales_qty_kg": 0.0,
                    "return_qty_kg": 0.0,
                    "net_sales_kg": 0.0,
                    "transaction_rows": 0,
                    "sales_rows": 0,
                    "return_rows": 0,
                }
                active_product_count = sum(
                    1
                    for product in catalog
                    if product["category_name"] == category["category_name"]
                    and (sales_date, product["product_code"]) in flow_agg
                )
                yield {
                    "sales_date": sales_date.isoformat(),
                    "date_has_records": date_has_records,
                    "category_code": category["category_code"],
                    "category_name": category["category_name"],
                    "category_has_records": category_has_records,
                    "sales_qty_kg": rounded(bucket["sales_qty_kg"]),
                    "return_qty_kg": rounded(bucket["return_qty_kg"]),
                    "net_sales_kg": rounded(bucket["net_sales_kg"]),
                    "transaction_rows": bucket["transaction_rows"],
                    "sales_rows": bucket["sales_rows"],
                    "return_rows": bucket["return_rows"],
                    "active_product_count": active_product_count,
                }

    write_csv(OUT_DIR / "daily_category_panel.csv", category_fields, category_rows())

    category_totals = {
        category["category_name"]: {
            "sales_qty_kg": 0.0,
            "return_qty_kg": 0.0,
            "net_sales_kg": 0.0,
        }
        for category in category_order
    }
    for (sales_date, category_name), bucket in category_agg.items():
        del sales_date
        for field in ("sales_qty_kg", "return_qty_kg", "net_sales_kg"):
            category_totals[category_name][field] += float(bucket[field])

    report = {
        "source": {
            "catalog": str(CATALOG_PATH),
            "flow": str(FLOW_PATH),
        },
        "catalog": {
            "product_count": len(catalog),
            "category_count": len(category_order),
            "category_product_counts": {
                category["category_name"]: sum(
                    1 for product in catalog if product["category_name"] == category["category_name"]
                )
                for category in category_order
            },
        },
        "flow": {
            "raw_rows": raw_rows,
            "sales_rows": raw_sales_rows,
            "return_rows": raw_return_rows,
            "observed_product_count": len(observed_codes),
            "unobserved_catalog_product_count": len(set(catalog_by_code) - observed_codes),
            "min_date": min_date.isoformat(),
            "max_date": max_date.isoformat(),
            "calendar_day_count": len(all_dates),
            "observed_day_count": len(observed_dates),
            "unobserved_day_count": len(set(all_dates) - observed_dates),
            "sales_qty_kg": rounded(raw_sales_qty),
            "return_qty_kg": rounded(raw_return_qty),
            "net_sales_kg": rounded(raw_sales_qty + raw_return_qty),
            "aggregated_item_key_count": len(flow_agg),
            "aggregated_category_key_count": len(category_agg),
        },
        "panels": {
            "daily_item_rows": len(all_dates) * len(catalog),
            "daily_category_rows": len(all_dates) * len(category_order),
            "missing_dates": sorted(
                sales_date.isoformat() for sales_date in set(all_dates) - observed_dates
            ),
        },
        "category_totals": {
            category_name: {field: rounded(value) for field, value in values.items()}
            for category_name, values in category_totals.items()
        },
    }
    (OUT_DIR / "panel_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
